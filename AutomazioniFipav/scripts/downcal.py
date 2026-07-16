import argparse
import html
import http.cookiejar
import re
import ssl
import sys
import time
import unicodedata
from datetime import date
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import HTTPCookieProcessor, HTTPSHandler, Request, build_opener, urlopen

from openpyxl import load_workbook


COMITATO = "09042"
BASE_URL = "https://fipavonline.it"
CHAMPIONSHIPS_URL = f"{BASE_URL}/main/tutti_i_campionati/{COMITATO}"
CALENDAR_URL = f"{BASE_URL}/gironi/stampa_calendario/{{girone_id}}"
MATCHES_URL = f"{BASE_URL}/main/gare_girone/{{girone_id}}"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
XLSX_MAGIC = b"PK\x03\x04"
SKIP_STATUSES = frozenset({"convertito", "scaricato"})
RETRYABLE_HTTP_CODES = {500, 502, 503, 504}


def make_opener():
    context = ssl._create_unverified_context()
    cookie_jar = http.cookiejar.CookieJar()
    return build_opener(HTTPSHandler(context=context), HTTPCookieProcessor(cookie_jar))


def fetch_bytes(url: str, opener=None) -> bytes:
    context = ssl._create_unverified_context()
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    if opener:
        with opener.open(request, timeout=30) as response:
            return response.read()
    with urlopen(request, context=context, timeout=30) as response:
        return response.read()


def is_retryable_error(exc: BaseException) -> bool:
    if isinstance(exc, HTTPError):
        return exc.code in RETRYABLE_HTTP_CODES
    return isinstance(exc, (TimeoutError, URLError))


def fetch_bytes_with_retry(url: str, opener=None, max_retries: int = 3, backoff: float = 2.0) -> bytes:
    last_exc: BaseException | None = None
    for attempt in range(max_retries):
        try:
            return fetch_bytes(url, opener=opener)
        except BaseException as exc:
            last_exc = exc
            if not is_retryable_error(exc) or attempt == max_retries - 1:
                raise
            time.sleep(backoff * (attempt + 1))
    raise last_exc  # pragma: no cover


def is_xlsx(data: bytes) -> bool:
    return data.startswith(XLSX_MAGIC)


def clean_text(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value)
    value = html.unescape(value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii")
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "da_verificare"


def parse_gironi(page_html: str) -> list[dict[str, str]]:
    sections = re.findall(r"(?is)<section\b[^>]*>(.*?)</section>", page_html)
    gironi = []
    seen = set()

    for section in sections:
        campionato_match = re.search(r'(?is)<h2\b[^>]*id="edit_torneo"[^>]*>(.*?)</h2>', section)
        campionato = clean_text(campionato_match.group(1)) if campionato_match else "DA_VERIFICARE"

        item_pattern = re.compile(
            r'(?is)<li\b[^>]*class="[^"]*\bspan4\b[^"]*"[^>]*>.*?'
            r'href="/(?:gironi/edit|main/gare_girone)/(\d+)".*?'
            r'<h4\b[^>]*>(.*?)</h4>.*?</li>'
        )
        for item in item_pattern.finditer(section):
            girone_id = item.group(1)
            if girone_id in seen:
                continue
            seen.add(girone_id)

            title = clean_text(item.group(2))
            girone = title
            if campionato != "DA_VERIFICARE":
                girone = re.sub(re.escape(campionato), "", title, count=1).strip()
                girone = re.sub(r"^\s*-\s*", "", girone).strip()
            if not girone:
                girone = "DA_VERIFICARE"

            gironi.append(
                {
                    "id": girone_id,
                    "campionato": campionato,
                    "girone": girone,
                    "titolo": title,
                }
            )

    return gironi


def read_registry_status(registry_path: Path) -> dict[str, str]:
    if not registry_path.exists():
        return {}
    status_by_id: dict[str, str] = {}
    for line in registry_path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("|") or line.startswith("| ---"):
            continue
        parts = [part.strip() for part in line.strip("|").split("|")]
        if len(parts) >= 5 and parts[4].isdigit():
            status_by_id[parts[4]] = parts[0].lower()
    return status_by_id


def has_unverified_names(meta: dict[str, str]) -> bool:
    return meta["campionato"] == "DA_VERIFICARE" or meta["girone"] == "DA_VERIFICARE"


def resolve_status_and_note(matches: int, note: str, meta: dict[str, str]) -> tuple[str, str]:
    if matches == 0:
        return "da_verificare", "nessuna gara pubblicata"
    if has_unverified_names(meta):
        return "convertito", f"{note}; nome girone DA_VERIFICARE"
    return "convertito", note


def cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_has_values(values: list[str]) -> bool:
    return any(value for value in values)


def markdown_table(headers: list[str], rows: list[list[str]]) -> list[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        escaped = [value.replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return lines


def convert_xlsx_to_markdown(xlsx_path: Path, md_path: Path, meta: dict[str, str]) -> int:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook["Calendario"] if "Calendario" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    title = cell_to_text(sheet.cell(row=1, column=1).value) or meta["titolo"]
    lines = [
        f"# {title}",
        "",
        f"- Campionato: {meta['campionato']}",
        f"- Girone: {meta['girone']}",
        f"- Numero girone: {meta['id']}",
        f"- Fonte: {CALENDAR_URL.format(girone_id=meta['id'])}",
        "",
    ]

    current_day = None
    current_rows = []
    headers = [
        "Codice gara",
        "Giorno",
        "Data",
        "Ora",
        "Impianto",
        "Comune",
        "Squadra casa",
        "Squadra ospite",
    ]
    total_matches = 0

    def flush_day() -> None:
        nonlocal total_matches
        if current_day and current_rows:
            lines.append(f"## {current_day}")
            lines.append("")
            lines.extend(markdown_table(headers, current_rows))
            lines.append("")
            total_matches += len(current_rows)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [cell_to_text(value) for value in row[:8]]
        if not row_has_values(values):
            continue
        first = values[0]
        if first.lower().startswith("giornata"):
            flush_day()
            current_day = first
            current_rows = []
            continue
        if first.isdigit() and len(values) >= 8:
            current_rows.append(values)

    flush_day()

    md_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return total_matches


def extract_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return clean_text(match.group(1)) if match else ""


def split_date_time(value: str) -> tuple[str, str, str]:
    parts = value.split()
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return "", parts[0], parts[1]
    return "", value, ""


def convert_matches_html_to_markdown(html_text: str, md_path: Path, meta: dict[str, str]) -> int:
    title = meta["titolo"]
    lines = [
        f"# {title}",
        "",
        f"- Campionato: {meta['campionato']}",
        f"- Girone: {meta['girone']}",
        f"- Numero girone: {meta['id']}",
        f"- Fonte: {MATCHES_URL.format(girone_id=meta['id'])}",
        "",
    ]

    headers = [
        "Codice gara",
        "Giorno",
        "Data",
        "Ora",
        "Impianto",
        "Comune",
        "Squadra casa",
        "Squadra ospite",
        "Risultato",
        "Parziali",
    ]
    total_matches = 0
    sections = re.findall(
        r'(?is)<div class="gare-wrap" id="g_\d+">(.*?)(?=<div class="gare-wrap" id="g_\d+">|</div>\s*<script|\Z)',
        html_text,
    )

    for section in sections:
        day_title = extract_first(r'<span class="h3-wrap">(.*?)</span>', section)
        if "/" in day_title:
            day_title = day_title.split("/")[-1].strip()
        if not day_title:
            day_title = "Giornata DA_VERIFICARE"

        rows = []
        blocks = re.findall(
            r'(?is)<div class="risultati" id="gara_\d+">(.*?)(?=<div class="risultati" id="gara_\d+">|</div>\s*</div>\s*<div class="gare-wrap"|</div>\s*</div>\s*</div>\s*<script|\Z)',
            section,
        )
        for block in blocks:
            data_gara = extract_first(r'<div class="info-gara-data">(.*?)</div>', block)
            giorno, data, ora = split_date_time(data_gara)
            codice = extract_first(r'<div class="info-gara-giornata">\s*Gara\s*([^<]+)</div>', block)
            impianto = extract_first(r'<div class="info-gara-campo-desc">(.*?)</div>', block)
            comune = extract_first(r'<div class="info-gara-campo-loc">(.*?)</div>', block)
            teams = re.findall(r'(?is)<span class="sq-nLong">(.*?)</span>', block)
            casa = clean_text(teams[0]) if len(teams) > 0 else ""
            ospite = clean_text(teams[1]) if len(teams) > 1 else ""
            risultato = extract_first(r'<span class="s-scoreText risultato-ufficiale">(.*?)</span>', block)
            parziali = extract_first(r'<span class="text-parziali">(.*?)</span>', block)

            if codice or casa or ospite:
                rows.append([codice, giorno, data, ora, impianto, comune, casa, ospite, risultato, parziali])

        if rows:
            lines.append(f"## {day_title}")
            lines.append("")
            lines.extend(markdown_table(headers, rows))
            lines.append("")
            total_matches += len(rows)

    md_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return total_matches


def fetch_and_convert_html(
    girone_id: str,
    html_path: Path,
    md_path: Path,
    girone: dict[str, str],
    opener,
    fallback_reason: str,
) -> tuple[int, str]:
    html_data = fetch_bytes_with_retry(MATCHES_URL.format(girone_id=girone_id), opener=opener).decode(
        "utf-8", "replace"
    )
    html_path.write_text(html_data, encoding="utf-8")
    matches = convert_matches_html_to_markdown(html_data, md_path, girone)
    note = f"{matches} gare da HTML; {fallback_reason}"
    return matches, note


def append_registry_row(registry_path: Path, row: list[str]) -> None:
    with registry_path.open("a", encoding="utf-8", newline="\n") as handle:
        handle.write("| " + " | ".join(value.replace("|", "\\|") for value in row) + " |\n")


def registry_template(season: str) -> str:
    return "\n".join(
        [
            f"# Calendari {season}",
            "",
            f"Registro dei calendari scaricati dal sito FIPAV Online per la stagione `{season}`.",
            "",
            "```text",
            f"Codice comitato: {COMITATO}",
            f"Pagina campionati: {CHAMPIONSHIPS_URL}",
            "```",
            "",
            "| Stato | Data download | Campionato | Girone | Numero girone | File input | File output | Note |",
            "| --- | --- | --- | --- | --- | --- | --- | --- |",
            "",
        ]
    )


def ensure_registry(registry_path: Path, season: str, reset: bool = False) -> None:
    if registry_path.exists() and not reset:
        return
    registry_path.parent.mkdir(parents=True, exist_ok=True)
    registry_path.write_text(registry_template(season), encoding="utf-8")


def process_girone(
    girone: dict[str, str],
    input_dir: Path,
    output_dir: Path,
    season_dir: Path,
    opener,
) -> tuple[str, str, str, str]:
    girone_id = girone["id"]
    basename = (
        f"girone_{girone_id}_"
        f"{slugify(girone['campionato'])}_"
        f"{slugify(girone['girone'])}"
    )
    xlsx_path = input_dir / f"{basename}.xlsx"
    html_path = input_dir / f"{basename}.html"
    md_path = output_dir / f"{basename}.md"

    data = fetch_bytes_with_retry(CALENDAR_URL.format(girone_id=girone_id), opener=opener)

    if is_xlsx(data):
        xlsx_path.write_bytes(data)
        matches = convert_xlsx_to_markdown(xlsx_path, md_path, girone)
        note = f"{matches} gare da XLSX"
        if matches == 0:
            matches, note = fetch_and_convert_html(
                girone_id, html_path, md_path, girone, opener, "XLSX senza righe gara"
            )
    else:
        html_path.write_bytes(data)
        matches, note = fetch_and_convert_html(
            girone_id, html_path, md_path, girone, opener, "risposta non-XLSX"
        )

    status, note = resolve_status_and_note(matches, note, girone)
    input_ref = xlsx_path.relative_to(season_dir).as_posix()
    if not is_xlsx(data) and html_path.exists():
        input_ref = html_path.relative_to(season_dir).as_posix()

    return status, note, input_ref, md_path.relative_to(season_dir).as_posix()


def run(season: str, force: bool, rebuild_registry: bool) -> int:
    root = Path.cwd()
    season_dir = root / "stagioni" / season
    input_dir = season_dir / "data" / "input" / "calendari"
    output_dir = season_dir / "data" / "output" / "calendari"
    registry_path = season_dir / "calendari.md"

    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    ensure_registry(registry_path, season, reset=rebuild_registry)

    opener = make_opener()
    page = fetch_bytes_with_retry(CHAMPIONSHIPS_URL, opener=opener).decode("utf-8", "replace")
    gironi = parse_gironi(page)
    registry_status = read_registry_status(registry_path)

    if not gironi:
        print("ERRORE: nessun girone trovato nella pagina campionati — probabile cambio layout HTML.", file=sys.stderr)
        return 1

    print(f"Trovati {len(gironi)} gironi nella pagina campionati.")
    print(f"Gia presenti nel registro: {len(registry_status)}.")

    today = date.today().isoformat()
    processed = 0
    skipped = 0
    errors = 0
    to_verify = 0
    verify_ids: list[str] = []
    error_ids: list[str] = []

    for girone in gironi:
        girone_id = girone["id"]
        if not force:
            status = registry_status.get(girone_id)
            if status in SKIP_STATUSES:
                skipped += 1
                continue

        try:
            status, note, input_ref, output_ref = process_girone(
                girone, input_dir, output_dir, season_dir, opener
            )
            append_registry_row(
                registry_path,
                [
                    status,
                    today,
                    girone["campionato"],
                    girone["girone"],
                    girone_id,
                    input_ref,
                    output_ref,
                    note,
                ],
            )
            processed += 1
            registry_status[girone_id] = status

            if status == "da_verificare":
                to_verify += 1
                verify_ids.append(girone_id)
                print(f"DA_VERIFICARE {girone_id} - {girone['titolo']} ({note})")
            else:
                print(f"OK {girone_id} - {girone['titolo']} ({note})")
        except Exception as exc:
            errors += 1
            error_ids.append(girone_id)
            basename = (
                f"girone_{girone_id}_"
                f"{slugify(girone['campionato'])}_"
                f"{slugify(girone['girone'])}"
            )
            append_registry_row(
                registry_path,
                [
                    "errore",
                    today,
                    girone["campionato"],
                    girone["girone"],
                    girone_id,
                    (input_dir / f"{basename}.xlsx").relative_to(season_dir).as_posix(),
                    (output_dir / f"{basename}.md").relative_to(season_dir).as_posix(),
                    str(exc).replace("\n", " ")[:180],
                ],
            )
            registry_status[girone_id] = "errore"
            print(f"ERRORE {girone_id} - {girone['titolo']}: {exc}", file=sys.stderr)

    print(
        f"Completato: {processed} convertiti, {skipped} saltati, {errors} errori, {to_verify} da_verificare."
    )
    if verify_ids:
        print(f"Elenco da_verificare: {', '.join(verify_ids)}")
    if error_ids:
        print(f"Elenco errori: {', '.join(error_ids)}")

    return 1 if errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Scarica e converte i calendari FIPAV Online.")
    parser.add_argument("--season", default="2025_2026", help="Stagione in formato YYYY_YYYY.")
    parser.add_argument("--force", action="store_true", help="Riscarica anche i gironi gia presenti nel registro.")
    parser.add_argument("--rebuild-registry", action="store_true", help="Ricostruisce da zero calendari.md.")
    args = parser.parse_args()
    return run(args.season, args.force, args.rebuild_registry)


if __name__ == "__main__":
    raise SystemExit(main())
